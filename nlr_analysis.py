import pandas as pd
import numpy as np
from scipy.optimize import curve_fit
from scipy.stats import sem
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
import warnings
warnings.filterwarnings('ignore')

FILE = //add file location
OUT  = //add file location
IMG  = //add file location

# ── Load & parse ───────────────────────────────────────────────────────────────
df_raw = pd.read_excel(FILE, header=None)
header_row = df_raw.iloc[0]
cond_cols = {}
for c, val in enumerate(header_row):
    if str(val).strip() in ['WT', 'KO', 'Alpha', 'Beta', 'Gamma']:
        cond_cols[str(val).strip()] = c

time_pts = [int(df_raw.iloc[r, 1]) for r in range(1, 6)]
conditions = ['WT', 'KO', 'Alpha', 'Beta', 'Gamma']
cond_order = sorted(cond_cols.items(), key=lambda x: x[1])
cond_ranges = {}
for i, (cname, sc) in enumerate(cond_order):
    end_c = cond_order[i+1][1] if i+1 < len(cond_order) else df_raw.shape[1]
    cond_ranges[cname] = (sc + 1, end_c)

def get_values(cname, ti):
    sc, ec = cond_ranges[cname]
    row = df_raw.iloc[ti + 1, sc:ec]
    return pd.to_numeric(row, errors='coerce').dropna().values

summary = {}
for cname in conditions:
    means, sems_arr = [], []
    for ti in range(5):
        v = get_values(cname, ti)
        means.append(np.mean(v))
        sems_arr.append(sem(v))
    summary[cname] = {'mean': means, 'sem': sems_arr}

# ── Model definitions ──────────────────────────────────────────────────────────
def exp_single(x, y0, plateau, tau):
    return plateau + (y0 - plateau) * np.exp(-x / tau)

# Double exponential: plateau + fast component + slow component
# f = fraction in fast component (0-1)
def exp_double(x, plateau, A1, tau1, A2, tau2):
    return plateau + A1 * np.exp(-x / tau1) + A2 * np.exp(-x / tau2)

x_fit   = np.array(time_pts, dtype=float)
x_dense = np.linspace(30, 1100, 500)
fit_to_run = [c for c in conditions if c != 'KO']

fit_results = {}
fit_curves  = {}

for cname in fit_to_run:
    y = np.array(summary[cname]['mean'])
    w = 1.0 / np.array(summary[cname]['sem'])

    # Single exponential (baseline for all)
    popt1, pcov1 = curve_fit(exp_single, x_fit, y, p0=[y[0], y[-1], 200.0],
                             bounds=([0,0,1],[10,10,5000]),
                             sigma=1/w, absolute_sigma=False, maxfev=10000)
    perr1 = np.sqrt(np.diag(pcov1))
    y_pred1 = exp_single(x_fit, *popt1)
    r2_single = 1 - np.sum((y - y_pred1)**2) / np.sum((y - np.mean(y))**2)

    if cname == 'Gamma':
        # Double exponential for Gamma
        # p0: plateau, A1 (fast amp), tau1 (fast), A2 (slow amp), tau2 (slow)
        # tau1 fixed at 50 ms (shortest measurable duration).
        # Fit the remaining 4 parameters: plateau, A1, A2, tau2.
        TAU1_FIXED = 50.0
        def exp_double_fixed_tau1(x, plateau, A1, A2, tau2):
            return plateau + A1 * np.exp(-x / TAU1_FIXED) + A2 * np.exp(-x / tau2)

        p0_d = [y[-1], (y[0]-y[-1])*0.6, (y[0]-y[-1])*0.4, 600.0]
        bounds_d = ([0,  0,  0,  100],
                    [5, 10, 10, 5000])
        popt2, pcov2 = curve_fit(exp_double_fixed_tau1, x_fit, y, p0=p0_d, bounds=bounds_d,
                                 sigma=1/w, absolute_sigma=False, maxfev=20000)
        perr2 = np.sqrt(np.diag(pcov2))
        y_pred2 = exp_double_fixed_tau1(x_fit, *popt2)
        r2_double = 1 - np.sum((y - y_pred2)**2) / np.sum((y - np.mean(y))**2)

        print("Gamma single:  y0={:.3f}, plateau={:.3f}, tau={:.1f}  R2={:.4f}".format(
            popt1[0], popt1[1], popt1[2], r2_single))
        print("Gamma double (tau1 fixed=50ms):  plateau={:.3f}+-{:.3f}, A1={:.3f}+-{:.3f}, A2={:.3f}+-{:.3f}, tau2={:.1f}+-{:.1f}  R2={:.4f}".format(
            popt2[0], perr2[0], popt2[1], perr2[1], popt2[2], perr2[2], popt2[3], perr2[3], r2_double))

        fit_results[cname] = {
            'model': 'double',
            'plateau': popt2[0], 'plateau_err': perr2[0],
            'A1': popt2[1], 'A1_err': perr2[1],
            'tau1': TAU1_FIXED,  'tau1_err': None,   # fixed — no SE
            'A2': popt2[2], 'A2_err': perr2[2],
            'tau2': popt2[3], 'tau2_err': perr2[3],
            'R2': r2_double,
            'R2_single': r2_single,
        }
        fit_curves[cname] = exp_double_fixed_tau1(x_dense, *popt2)
    else:
        print("{}: y0={:.3f}, plateau={:.3f}, tau={:.1f}  R2={:.4f}".format(
            cname, popt1[0], popt1[1], popt1[2], r2_single))
        fit_results[cname] = {
            'model': 'single',
            'y0': popt1[0], 'y0_err': perr1[0],
            'plateau': popt1[1], 'plateau_err': perr1[1],
            'tau': popt1[2], 'tau_err': perr1[2],
            'R2': r2_single,
        }
        fit_curves[cname] = exp_single(x_dense, *popt1)

# ── Plot ───────────────────────────────────────────────────────────────────────
colors  = {'WT': '#1f77b4', 'KO': '#888888', 'Alpha': '#2ca02c', 'Beta': '#ff7f0e', 'Gamma': '#d62728'}
markers = {'WT': 'o', 'KO': 's', 'Alpha': '^', 'Beta': 'D', 'Gamma': 'v'}

fig, ax = plt.subplots(figsize=(8.0, 5.5))

# KO: scatter + dashed line, no fit
ax.errorbar(time_pts, summary['KO']['mean'], yerr=summary['KO']['sem'],
            fmt=markers['KO'], color=colors['KO'], markersize=6,
            elinewidth=1.5, capsize=4, capthick=1.5, linestyle='--', label='KO (no fit)', alpha=0.65)

for cname in fit_to_run:
    fr = fit_results[cname]
    ax.errorbar(time_pts, summary[cname]['mean'], yerr=summary[cname]['sem'],
                fmt=markers[cname], color=colors[cname], markersize=6,
                linestyle='none', elinewidth=1.5, capsize=4, capthick=1.5, zorder=5)
    if fr['model'] == 'single':
        lbl = "{} [y0={:.2f}, plat={:.2f}, τ={:.0f} ms, R²={:.3f}]".format(
            cname, fr['y0'], fr['plateau'], fr['tau'], fr['R2'])
    else:
        lbl = ("{} [plat={:.2f}, A1={:.2f} τ₁=50ms(fixed), "
               "A2={:.2f} τ₂={:.0f} ms, R²={:.3f}]").format(
            cname, fr['plateau'], fr['A1'], fr['A2'], fr['tau2'], fr['R2'])
    ax.plot(x_dense, fit_curves[cname], color=colors[cname], linewidth=2.0, label=lbl, zorder=4)

ax.set_xscale('log')
ax.set_xticks(time_pts)
ax.get_xaxis().set_major_formatter(ticker.ScalarFormatter())
ax.set_xlabel('Stimulus Duration (ms)', fontsize=12)
ax.set_ylabel('Response Ratio', fontsize=12)
ax.set_title('Tuning Curves  |  Single exp (WT/Alpha/Beta)  +  Double exp (Gamma)\n(KO excluded from fitting)',
             fontsize=11, fontweight='bold')
ax.legend(fontsize=7.2, loc='upper right', framealpha=0.92)
ax.axhline(1.0, color='black', linewidth=0.7, linestyle=':', alpha=0.5)
ax.grid(True, which='both', alpha=0.25, linewidth=0.5)
ax.spines[['top','right']].set_visible(False)
plt.tight_layout()
plt.savefig(IMG, dpi=180, bbox_inches='tight')
plt.close()
print('Chart saved:', IMG)

# ── Excel output ───────────────────────────────────────────────────────────────
hdr_fill  = PatternFill("solid", fgColor="2F5496")
hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
sub_fill  = PatternFill("solid", fgColor="D9E1F2")
sub_font  = Font(bold=True, name="Arial", size=10)
body_font = Font(name="Arial", size=10)
bold_font = Font(bold=True, name="Arial", size=10)
warn_fill = PatternFill("solid", fgColor="FCE4D6")
center    = Alignment(horizontal="center", vertical="center")
thin      = Side(style="thin")
bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

def hc(ws, r, c, v):
    cell = ws.cell(r, c, v); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = center; cell.border = bdr

def sc(ws, r, c, v):
    cell = ws.cell(r, c, v); cell.fill = sub_fill; cell.font = sub_font
    cell.alignment = center; cell.border = bdr

def bc(ws, r, c, v, fmt=None, bold=False, warn=False):
    cell = ws.cell(r, c, v)
    cell.font = bold_font if bold else body_font
    cell.alignment = center; cell.border = bdr
    if fmt: cell.number_format = fmt
    if warn: cell.fill = warn_fill

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Summary & Fit"

# Title
ws1.merge_cells("A1:J1")
t = ws1.cell(1, 1, "Non-Linear Regression Results  |  Single exp: y = plateau + (y0-plateau)*exp(-x/tau)  |  Double exp (Gamma): y = plateau + A1*exp(-x/tau1) + A2*exp(-x/tau2)")
t.font = Font(bold=True, name="Arial", size=11)
t.alignment = Alignment(horizontal="center")
ws1.row_dimensions[1].height = 20

# ── Mean ± SEM table ──
row = 3
hc(ws1, row, 1, "Time (ms)")
col_map = {}
c = 2
for cname in conditions:
    ws1.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c+1)
    hc(ws1, row, c, cname + ("  [excluded]" if cname == "KO" else ""))
    col_map[cname] = c; c += 2

row += 1
for cname in conditions:
    sc(ws1, row, col_map[cname],   "Mean")
    sc(ws1, row, col_map[cname]+1, "SEM")
sc(ws1, row, 1, "Time (ms)")

for ti, t_val in enumerate(time_pts):
    row += 1
    bc(ws1, row, 1, t_val, "#,##0")
    for cname in conditions:
        bc(ws1, row, col_map[cname],   round(summary[cname]["mean"][ti], 4), "0.0000")
        bc(ws1, row, col_map[cname]+1, round(summary[cname]["sem"][ti],  4), "0.0000")

def safe_round(v, n):
    """Return rounded float, or None if inf/nan."""
    try:
        if v is None or (isinstance(v, float) and (np.isinf(v) or np.isnan(v))):
            return None
        return round(float(v), n)
    except Exception:
        return None

def bc_val(ws, r, c, val, fmt=None, bold=False, warn=False):
    """Write value; if None write 'N/A'."""
    if val is None:
        cell = ws.cell(r, c, "N/A")
        cell.font = Font(italic=True, color="FF0000", name="Arial", size=10)
        cell.alignment = center; cell.border = bdr
    else:
        bc(ws, r, c, val, fmt, bold, warn)

# ── Fit parameters: single-exp conditions ──
row += 2
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
ws1.cell(row, 1, "Single Exponential Fit  [WT, Alpha, Beta]  —  y = plateau + (y0 - plateau) * exp(-x / tau)").font = Font(bold=True, name="Arial", size=11)
row += 1
for lbl, ci in zip(["Condition", "y0", "± SE (y0)", "Plateau", "± SE (Plateau)", "tau (ms)", "± SE (tau)", "R²"], range(1, 9)):
    hc(ws1, row, ci, lbl)
for cname in [c for c in fit_to_run if fit_results[c]['model'] == 'single']:
    row += 1
    fr = fit_results[cname]
    bc(ws1, row, 1, cname, bold=True)
    bc_val(ws1, row, 2, safe_round(fr["y0"],           4), "0.0000")
    bc_val(ws1, row, 3, safe_round(fr["y0_err"],        4), "0.0000")
    bc_val(ws1, row, 4, safe_round(fr["plateau"],       4), "0.0000")
    bc_val(ws1, row, 5, safe_round(fr["plateau_err"],   4), "0.0000")
    bc_val(ws1, row, 6, safe_round(fr["tau"],           2), "0.00")
    bc_val(ws1, row, 7, safe_round(fr["tau_err"],       2), "0.00")
    bc_val(ws1, row, 8, safe_round(fr["R2"],            4), "0.0000")

# ── Fit parameters: double-exp Gamma ──
row += 2
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
ws1.cell(row, 1, "Double Exponential Fit  [Gamma]  —  y = plateau + A1*exp(-x/tau1) + A2*exp(-x/tau2)").font = Font(bold=True, name="Arial", size=11)
row += 1
for lbl, ci in zip(["Condition", "Plateau", "± SE", "A1 (fast)", "± SE",
                     "tau1 (ms)", "± SE", "A2 (slow)", "± SE", "tau2 (ms)", "± SE", "R²"], range(1, 13)):
    hc(ws1, row, ci, lbl)
row += 1
fr = fit_results['Gamma']
bc(ws1, row, 1,  "Gamma", bold=True)
bc_val(ws1, row, 2,  safe_round(fr["plateau"],      4), "0.0000")
bc_val(ws1, row, 3,  safe_round(fr["plateau_err"],  4), "0.0000")
bc_val(ws1, row, 4,  safe_round(fr["A1"],           4), "0.0000")
bc_val(ws1, row, 5,  safe_round(fr["A1_err"],       4), "0.0000")
bc_val(ws1, row, 6,  safe_round(fr["tau1"],         2), "0.00", warn=True)   # fixed param
cell_fixed = ws1.cell(row, 7, "fixed")
cell_fixed.font = Font(italic=True, color="7F6000", name="Arial", size=10)
cell_fixed.fill = PatternFill("solid", fgColor="FFF2CC")
cell_fixed.alignment = center; cell_fixed.border = bdr
bc_val(ws1, row, 8,  safe_round(fr["A2"],           4), "0.0000")
bc_val(ws1, row, 9,  safe_round(fr["A2_err"],       4), "0.0000")
bc_val(ws1, row, 10, safe_round(fr["tau2"],         2), "0.00")
bc_val(ws1, row, 11, safe_round(fr["tau2_err"],     2), "0.00")
bc_val(ws1, row, 12, safe_round(fr["R2"],           4), "0.0000")

# R² comparison for Gamma
row += 2
ws1.cell(row, 1, "Gamma R² comparison:").font = Font(bold=True, name="Arial", size=10)
bc(ws1, row, 2, "Single exp R²", bold=True)
bc_val(ws1, row, 3, safe_round(fr["R2_single"], 4), "0.0000", warn=True)
bc(ws1, row, 4, "Double exp R²", bold=True)
bc_val(ws1, row, 5, safe_round(fr["R2"],        4), "0.0000")

# Explanatory note for N/A SE
row += 2
note_fill = PatternFill("solid", fgColor="FFF2CC")
note_text = (
    "Note (Gamma — τ₁ fixed): The fast time constant τ₁ was held fixed at 50 ms (the shortest "
    "measured stimulus duration) because the free-fitting optimizer consistently converged to τ₁ < 50 ms, "
    "outside the measurement window, making τ₁ non-identifiable. With τ₁ fixed, the remaining four "
    "parameters (plateau, A₁, A₂, τ₂) were fitted freely and SE is fully reportable for those. "
    "The τ₁ = 50 ms cell is shaded to indicate it is a fixed constraint, not a fitted value."
)
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
note_cell = ws1.cell(row, 1, note_text)
note_cell.fill = note_fill
note_cell.font = Font(italic=True, name="Arial", size=9, color="7F6000")
note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws1.row_dimensions[row].height = 48

col_widths = {"A": 14, "B": 12, "C": 14, "D": 12, "E": 14,
              "F": 12, "G": 14, "H": 12, "I": 14, "J": 12, "K": 14, "L": 12}
for col, w in col_widths.items():
    ws1.column_dimensions[col].width = w

# ── Fit Curve Data sheet ──
ws2 = wb.create_sheet("Fit Curve Data")
hc(ws2, 1, 1, "x (ms)")
c = 2
curve_col = {}
for cname in fit_to_run:
    hc(ws2, 1, c, cname + (" (double exp)" if cname == "Gamma" else " (single exp)"))
    curve_col[cname] = c; c += 1

x_dense_list = list(np.linspace(30, 1100, 500))
for i, xv in enumerate(x_dense_list):
    r = i + 2
    ws2.cell(r, 1, round(xv, 2))
    for cname in fit_to_run:
        ws2.cell(r, curve_col[cname], round(float(fit_curves[cname][i]), 5))

for col in ["A","B","C","D","E"]:
    ws2.column_dimensions[col].width = 18

# ── Raw Data sheet ──
ws3 = wb.create_sheet("Raw Data")
hc(ws3, 1, 1, "Condition"); hc(ws3, 1, 2, "Time (ms)"); hc(ws3, 1, 3, "Replicates")
row = 2
for cname in conditions:
    for ti, t_val in enumerate(time_pts):
        vals = get_values(cname, ti)
        ws3.cell(row, 1, cname); ws3.cell(row, 2, t_val)
        for j, v in enumerate(vals):
            ws3.cell(row, 3 + j, round(float(v), 4))
        row += 1

# ── Chart sheet ──
ws_chart = wb.create_sheet("Chart", 0)
xlimg = XLImage(IMG)
xlimg.anchor = 'B2'; xlimg.width = 740; xlimg.height = 520
ws_chart.add_image(xlimg)
ws_chart.column_dimensions['A'].width = 3

wb.save(OUT)
print('Excel saved:', OUT)
